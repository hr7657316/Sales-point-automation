"""Monthly point workbook in the shape of Allissa's rep worksheets.

One SUMMARY tab, one tab per rep (category quantities x rates with live
formulas, bonuses, splits, honorarium placeholder, new-customer
candidates), an ALL ROWS audit trail and a REVIEW tab. Read-only on all
inputs; writes a fresh .xlsx only.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .comp import load_comp_plans

FONT = "Arial"
BOLD = Font(name=FONT, bold=True)
NORMAL = Font(name=FONT)
BLUE = Font(name=FONT, color="0000FF")
HEAD_FILL = PatternFill("solid", fgColor="D9E1F2")

# rep code -> (worksheet name, comp-plan key)
REP_NAMES = {
    "M1-11-69": ("Paul Lopiccolo", "LOPICCOLO"),
    "M1-21-08": ("Austin Liput", "LIPUT"),
    "M1-11-3": ("Zach Schneider", None),
    "M1-11-31": ("Samir Thapa", None),
    "M1-11-38": ("Brendon Hink", None),
    "M1-21-4": ("Taylor Miller", None),
    "M1-19-20": ("Matthew Christensen", None),
    "M1-22-1": ("Jon Janerich", None),
    "M1-21-8": ("Logan Yearego", None),
    "M1-21-21": ("Pete Crosby", None),
    "M1-21-0": ("House East", None),
    "M1-21-2": ("House West", None),
    "M1-21-1": ("M1-21-1 (unmapped)", None),
}

# rule id -> worksheet category label (Allissa's wording)
CATEGORY_LABELS = {
    "TCT_WC_SURGICAL": "1 TCT Work comp Surgical",
    "TCT_WC_POST_SURGICAL": "1 TCT WC Post-Surgical MD/DO/PA/NP/DPM (<30 Post-op)",
    "TCT_WC_NONSURG_TRICARE_USDL_COLD": "1 TCT WC Non-Surgical/Tri care-USDL SX/TT(hot or cold)",
    "TCT_NONSURG_WC_AUTO_LITIGATED": "1 TCT Non Surgical WC & Auto open or Litigated SX or non SX",
    "TCT_MICH_AUTO_NON_LITIGATED": "1 TCT Mich Auto SX or Non SX (NON-Litigated)",
    "TCT_RR_PI_SLIPFALL_PIP": "1 TCT Railroad / Slip & Fall / Personal Injury / Auto PIP",
    "MZ_WORK_COMP": "1 MZ Work Comp",
    "MZ_GOV_AND_COMMERCIAL": "1 MZ Medicare/Auto/Tricare-USDL/Medicaid/Commercial",
    "MZ_PA_MI_FL_AUTO": "1 MZ PA, MI, FL, Auto",
    "MZ_OH_WC_RENTAL_TRIAL": "1 MZ/TENS OH Work Comp 13 Month Rental / Trial",
    "MZ_RR_PI_SLIPFALL_PIP": "1 MZ Railroad / Slip & Fall / Personal Injury / Auto PIP",
    "SPORTZ": "1 SportZ",
    "BONE_STIM_OR_LASER": "1 Bone Stim or Lazer",
    "BACK_BRACE_WC_MEDICARE": "Back Brace WC or Medicare",
    "KNEE_SCOOTER": "Knee Scooter",
    "GARMENT_ONLY": "Garment Only",
    "ANC_TCT_WC_SURGICAL_POST_OP": "ANCILLARY: 1 TCT WC Surgical- Post OP (<30 day post op)",
    "ANC_TCT_WC_NON_SURGICAL": "ANCILLARY: 1 TCT WC Non-Surgical",
    "ANC_COLD_THERAPY": "ANCILLARY: 1 TT (Hot or Cold)",
    "ANC_MZ_WORK_COMP": "ANCILLARY: 1 MZ Work comp",
    "ANC_MZ_AUTO": "ANCILLARY: 1 MZ Auto",
}
ZERO_RULES = {"SUPPLY_ONLY", "SELF_PAY_ZERO", "NOT_FIT", "INSURANCE_NOT_ELIGIBLE",
              "BMV_OVERRIDE", "NO_RULE_MATCH"}


def _rep_key(rep) -> str:
    return (rep.rep_id or rep.name).strip().upper()


def _set(ws, ref, value, font=NORMAL, fmt=None, fill=None):
    cell = ws[ref]
    cell.value = value
    cell.font = font
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    return cell


def build_workbook(results: list, month_label: str, out_path: Path,
                   new_customer_candidates: dict | None = None,
                   comp_plans_path: str = "rules/comp_plans.csv") -> Path:
    plans = load_comp_plans(comp_plans_path)
    candidates = new_customer_candidates or {}

    # ---- gather per-rep data ---------------------------------------------
    per_rep: dict = defaultdict(lambda: {
        "full": Counter(), "split_points": 0, "split_rows": [],
        "gold": 0, "zero_rows": [], "rows": []})
    for res in results:
        shares = res.rep_allocations
        for rep, pts in shares:
            key = _rep_key(rep)
            bucket = per_rep[key]
            bucket["rows"].append((res, rep, pts))
            if res.rule_used in ZERO_RULES:
                bucket["zero_rows"].append(res)
                continue
            if len(shares) > 1:
                bucket["split_points"] += pts
                bucket["split_rows"].append((res, pts))
            else:
                bucket["full"][(res.rule_used, res.base_points)] += 1
            if any("GOLD" in b for b in res.bonuses_applied):
                bucket["gold"] += 1 if len(shares) == 1 else 0

    wb = Workbook()
    summary = wb.active
    summary.title = "SUMMARY"

    # ---- per-rep tabs ------------------------------------------------------
    rep_totals = {}
    ordered = sorted(per_rep.items(), key=lambda kv: -sum(p for _, _, p in kv[1]["rows"]))
    for key, data in ordered:
        name, plan_key = REP_NAMES.get(key, (key, None))
        ws = wb.create_sheet(title=name[:31])
        ws.column_dimensions["A"].width = 62
        for col in "BCDE":
            ws.column_dimensions[col].width = 14
        _set(ws, "A1", "MONTHLY QUOTA WORKSHEET (system-generated)", BOLD)
        _set(ws, "B1", f"({key})", BOLD)
        _set(ws, "A2", "Employee:", BOLD)
        _set(ws, "B2", name)
        _set(ws, "A3", "Current Month", BOLD)
        _set(ws, "B3", month_label)
        _set(ws, "A5", "DME (Venous)", BOLD, fill=HEAD_FILL)
        _set(ws, "B5", "Rate", BOLD, fill=HEAD_FILL)
        _set(ws, "C5", "Quantity", BOLD, fill=HEAD_FILL)
        _set(ws, "D5", "Total", BOLD, fill=HEAD_FILL)
        row = 6
        first_row = row
        for rule_id, label in CATEGORY_LABELS.items():
            matches = [(b, n) for (r, b), n in data["full"].items() if r == rule_id]
            if not matches:
                continue
            for base, n in sorted(matches):
                _set(ws, f"A{row}", label)
                _set(ws, f"B{row}", base, BLUE, "#,##0")
                _set(ws, f"C{row}", n, BLUE, "0")
                _set(ws, f"D{row}", f"=B{row}*C{row}", NORMAL, "#,##0")
                row += 1
        _set(ws, f"A{row}", "1 Gold Pair Patient Bonus* (not applicable for FFW/AMP MI)")
        _set(ws, f"B{row}", 50, BLUE, "#,##0")
        _set(ws, f"C{row}", data["gold"], BLUE, "0")
        _set(ws, f"D{row}", f"=B{row}*C{row}", NORMAL, "#,##0")
        row += 1
        _set(ws, f"A{row}", "1 Split Account Multiplier 50% (half of shared-account points)")
        _set(ws, f"B{row}", "Split")
        _set(ws, f"C{row}", len(data["split_rows"]), BLUE, "0")
        _set(ws, f"D{row}", data["split_points"], BLUE, "#,##0")
        ws[f"D{row}"].comment = Comment(
            "Rows whose REP column names two reps; this rep's half is listed here. "
            "See ALL ROWS for the patients.", "engine")
        row += 1
        _set(ws, f"A{row}", "1 TCT or MZ new customer bonus FIT COMPLETE (CANDIDATES - confirm)")
        cands = candidates.get(key, [])
        _set(ws, f"B{row}", 500, BLUE, "#,##0")
        _set(ws, f"C{row}", 0, BLUE, "0")
        _set(ws, f"D{row}", f"=B{row}*C{row}", NORMAL, "#,##0")
        ws[f"C{row}"].comment = Comment(
            "Enter the confirmed count. Candidates from the tracker/history check are "
            "listed at the bottom of this tab.", "engine")
        new_row = row
        row += 1
        _set(ws, f"A{row}", "Honorarium deduction (50% of payout in month of engagement) - enter")
        _set(ws, f"B{row}", "")
        _set(ws, f"C{row}", "")
        _set(ws, f"D{row}", 0, BLUE, "#,##0")
        hon_row = row
        row += 2
        _set(ws, f"A{row}", "TOTAL POINTS:", BOLD)
        _set(ws, f"D{row}", f"=SUM(D{first_row}:D{new_row})-D{hon_row}", BOLD, "#,##0")
        total_row = row
        rep_totals[key] = (name, f"'{ws.title}'!D{total_row}",
                           sum(p for _, _, p in data["rows"]), plan_key, data)
        row += 1
        if plan_key and plan_key in plans:
            plan = plans[plan_key]
            _set(ws, f"A{row}",
                 "Commission ($) - band payout + highest bonus tier (per rep comp table)",
                 BOLD)
            base_pts = sum(p for _, _, p in data["rows"])
            _set(ws, f"D{row}", plan.commission_for(base_pts), BOLD, "$#,##0")
            ws[f"D{row}"].comment = Comment(
                "Computed from rules/comp_plans.csv on the base points shown; "
                "re-run after honorarium/new-customer entries change.", "engine")
            row += 1
        row += 1
        _set(ws, f"A{row}",
             "New provider CANDIDATES this month (in this report, not in any Jan-Jul report)",
             BOLD)
        row += 1
        for prov, lr in cands:
            _set(ws, f"A{row}", prov)
            _set(ws, f"B{row}", f"last referral {lr}")
            row += 1
        if not cands:
            _set(ws, f"A{row}", "(none found)")
            row += 1
        row += 1
        _set(ws, f"A{row}", "Rows scoring 0 (with reason)", BOLD)
        row += 1
        for res in data["zero_rows"]:
            _set(ws, f"A{row}", f"{res.row.patient} - {res.row.product[:40]}")
            _set(ws, f"B{row}", res.rule_used)
            row += 1
        _set(ws, f"A{row + 1}",
             "Blue = input from the Fit Report; black = formula. "
             "Every row's reasoning is on the ALL ROWS tab.",
             Font(name=FONT, italic=True))

    # ---- summary -----------------------------------------------------------
    heads = ["Rep", "Code", "Base points (engine)", "Gold pairs", "Split rows",
             "New-customer candidates", "Rows scoring 0", "Commission $ (where table known)"]
    for i, h in enumerate(heads, start=1):
        _set(summary, f"{get_column_letter(i)}1", h, BOLD, fill=HEAD_FILL)
        summary.column_dimensions[get_column_letter(i)].width = 24
    r = 2
    ranked = sorted(rep_totals.items(), key=lambda kv: -kv[1][2])
    for key, (name, ref, pts, plan_key, data) in ranked:
        _set(summary, f"A{r}", name)
        _set(summary, f"B{r}", key)
        _set(summary, f"C{r}", f"={ref}", NORMAL, "#,##0")
        _set(summary, f"D{r}", data["gold"], NORMAL, "0")
        _set(summary, f"E{r}", len(data["split_rows"]), NORMAL, "0")
        _set(summary, f"F{r}", len(candidates.get(key, [])), NORMAL, "0")
        _set(summary, f"G{r}", len(data["zero_rows"]), NORMAL, "0")
        if plan_key and plan_key in plans:
            _set(summary, f"H{r}", plans[plan_key].commission_for(pts), NORMAL, "$#,##0")
        r += 1
    _set(summary, f"A{r + 1}",
         f"Source: {month_label} Fit Report, scored by the point engine; "
         "honorarium and new-customer counts to be entered per rep tab.",
         Font(name=FONT, italic=True))

    # ---- all rows ----------------------------------------------------------
    ws = wb.create_sheet("ALL ROWS")
    cols = ["Rep", "Patient", "Provider", "Product", "Type", "Insurance status",
            "Fit date", "DOS", "Rule", "Base points", "Bonus", "Rep share", "Explanation"]
    for i, h in enumerate(cols, start=1):
        _set(ws, f"{get_column_letter(i)}1", h, BOLD, fill=HEAD_FILL)
    r = 2
    for res in results:
        for rep, pts in res.rep_allocations:
            vals = [str(rep), res.row.patient, res.row.pro, res.row.product, res.row.type,
                    res.row.insurance_status, str(res.row.fit_date or ""), res.row.dos_code,
                    res.rule_used, res.base_points, res.bonus_points, pts, res.explanation]
            for i, v in enumerate(vals, start=1):
                _set(ws, f"{get_column_letter(i)}{r}", v)
            r += 1
    for i, w in enumerate([22, 28, 26, 34, 12, 20, 12, 10, 30, 10, 8, 10, 90], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ---- review ------------------------------------------------------------
    ws = wb.create_sheet("REVIEW")
    review_heads = ["Rep", "Patient", "Provider", "Product", "Type", "Why flagged"]
    for i, h in enumerate(review_heads, start=1):
        _set(ws, f"{get_column_letter(i)}1", h, BOLD, fill=HEAD_FILL)
        ws.column_dimensions[get_column_letter(i)].width = 30
    r = 2
    for res in results:
        if res.review_needed:
            for i, v in enumerate([res.row.rep, res.row.patient, res.row.pro, res.row.product,
                                   res.row.type, res.explanation], start=1):
                _set(ws, f"{get_column_letter(i)}{r}", v)
            r += 1

    for sheet in wb.worksheets:
        for row_cells in sheet.iter_rows():
            for c in row_cells:
                if c.font.name != FONT:
                    c.font = Font(name=FONT, bold=c.font.bold, italic=c.font.italic,
                                  color=c.font.color)
                wrap = sheet.title == "ALL ROWS" and c.column == 13
                c.alignment = Alignment(vertical="top", wrap_text=wrap)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
